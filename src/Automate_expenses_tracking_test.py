import os
import glob
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ─────────────── CONFIG ───────────────
input_dir = r'./Input'
mapping_file = r'./mapping_categories_test.xlsx'
output_file = r'./expenses_tracker_test.xlsx'
# ──────────────────────────────────────

# 1. Load keyword → category mapping
mapping_df = pd.read_excel(mapping_file, sheet_name='Mappatura')
mapping_df.columns = [c.strip().lower() for c in mapping_df.columns]
mapping_df.rename(columns={'categoria': 'category'}, inplace=True) 
mapping_df['keyword'] = mapping_df['keyword'].astype(str).str.strip().str.lower()
category_map = mapping_df.groupby('category')['keyword'].apply(list).to_dict()

fixed_categories = {
    'Rent', 'Wi-Fi', 'Phone Subscription',
    'Insurance', 'Insurance Savings',
    'Gym Subscription', 'Groceries'
}

def categorize(desc: str) -> str:
    dl = str(desc).lower()
    for cat, kws in category_map.items():
        for kw in kws:
            if kw in dl:
                return cat
    return 'Other'

# 2. Read all input files
df_list = []

for path in glob.glob(os.path.join(input_dir, '*')):
    fname = os.path.basename(path).lower()

    # ─── Amex ───
    if fname == 'amex_test.xlsx':
        df = pd.read_excel(path, sheet_name='Details', skiprows=6)
        df = df[['Date', 'Description', 'Amount']]
        df['Origin'] = 'Amex'
        df['Date'] = pd.to_datetime(df['Date'], format='%m/%d/%Y', errors='coerce')
        df['Description'] = df['Description'].fillna('').astype(str)

        # Convert Amount to numeric (keep only positive values = expenses)
        s = (df['Amount'].astype(str)
             .str.replace('.', '', regex=False)
             .str.replace(',', '.', regex=False))
        df['Amount'] = pd.to_numeric(s, errors='coerce').fillna(0)
        df = df[df['Amount'] >= 0]

        # Append to list
        df_list.append(df[['Date', 'Description', 'Amount', 'Origin']])

if not df_list:
    print("⚠️ No valid files found.")
    exit()

tx = pd.concat(df_list, ignore_index=True)
print(f"⦿ Imported rows: {len(tx)}")

# 3. Load historical master file for categorization
try:
    old_tx = pd.read_excel(output_file, sheet_name='Master', engine='openpyxl')
    old_tx['Date'] = pd.to_datetime(old_tx['Date'], errors='coerce').dt.date
    old_tx['Description'] = old_tx['Description'].astype(str).str.strip().str.lower()
    old_tx['Amount'] = pd.to_numeric(old_tx['Amount'], errors='coerce')
    old_tx['Origin'] = old_tx['Origin'].astype(str).str.strip()
except Exception as e:
    print(f"⚠️ Master file invalid or missing: {e}")
    old_tx = pd.DataFrame()

# 4. Categorization
if not old_tx.empty and 'category' in old_tx.columns:
    desc_to_cat = (
        old_tx.dropna(subset=['Description', 'category'])
              .drop_duplicates(subset='Description')
              .set_index('Description')['category']
              .to_dict()
    )
    tx['Category'] = tx['Description'].str.strip().str.lower().map(desc_to_cat)
else:
    tx['Category'] = None

tx['Category'] = tx.apply(
    lambda row: categorize(row['Description']) if pd.isna(row['Category']) else row['Category'],
    axis=1
)
tx['TypeTransaction'] = 'Expense'
tx['TypeExpense'] = tx['Category'].apply(lambda c: 'Fixed' if c in fixed_categories else 'Variable')

# 5. Merge with master
try:
    combined = pd.concat([old_tx, tx], ignore_index=True)
    combined.drop_duplicates(subset=['Date', 'Description', 'Amount', 'Origin'], inplace=True)
    print(f"⦿ Added {len(combined) - len(old_tx)} unique records")
except Exception as e:
    print(f"⚠️ Error creating combined file: {e}")
    combined = tx

# 6. Save to Excel
combined['Date'] = pd.to_datetime(combined['Date'], dayfirst=True, errors='coerce').dt.date
combined['Amount'] = pd.to_numeric(combined['Amount'], errors='coerce')

if os.path.exists(output_file):
    wb = load_workbook(output_file)
else:
    wb = Workbook(); wb.remove(wb.active)

if 'Master' in wb.sheetnames:
    del wb['Master']

ws = wb.create_sheet('Master')
for row in dataframe_to_rows(combined, index=False, header=True):
    ws.append(row)

for cell in ws.iter_rows(min_row=2, min_col=1, max_col=1):
    cell[0].number_format = 'DD/MM/YYYY'
for cell in ws.iter_rows(min_row=2, min_col=3, max_col=3):
    cell[0].number_format = '#,##0.00'

wb.save(output_file)
print(f"✅ Completed: {len(combined)} records saved in '{output_file}'")

# 7. Update mapping file with new Description → Category pairs
try:
    map_xlsx = pd.read_excel(mapping_file, sheet_name='Mappatura')
    map_xlsx.columns = [c.strip().lower() for c in map_xlsx.columns]
    map_xlsx['keyword'] = map_xlsx['keyword'].astype(str).str.strip().str.lower()

    new_map = tx[['Description', 'Category']].dropna().copy()
    new_map['keyword'] = new_map['Description'].str.strip().str.lower()
    new_map.rename(columns={'Category': 'category'}, inplace=True)
    new_map = new_map[['category', 'keyword']].drop_duplicates()

    merged = new_map.merge(map_xlsx, how='left', on=['keyword', 'category'], indicator=True)
    to_append = merged[merged['_merge'] == 'left_only'][['category', 'keyword']]

    if not to_append.empty:
        updated = pd.concat([map_xlsx[['category', 'keyword']], to_append], ignore_index=True)
        updated.drop_duplicates(inplace=True)

        with pd.ExcelWriter(mapping_file, engine='openpyxl', mode='w') as writer:
            updated.to_excel(writer, index=False, sheet_name='Mappatura')

        print(f"🆕 mapping_categories.xlsx updated: added {len(to_append)} new rows.")
    else:
        print("✅ No new mappings to add.")

except Exception as e:
    print(f"⚠️ Error updating mapping_categories.xlsx: {e}")
